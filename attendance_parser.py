import re

import openpyxl
import pandas as pd

from attendance_generator import (
    capitalize_first_word_if_english,
    clean_name,
    format_text,
)


SHEET_CONFIGS = [
    ("영어", 6, None, "과정"),
    ("일본어", 7, 5, "과정.1"),
    ("중국어", 6, None, "과정1"),
    ("한국어", 6, None, "과정1"),
]

SUMMARY_MARKERS = {
    "실적",
    "목표",
    "달성율",
    "총원",
    "총계",
    "합계",
    "수강생",
    "신규",
    "휴학자",
}


def extract_duration(text):
    if not text:
        return None
    for line in reversed(text.strip().splitlines()):
        if any(token in line for token in ["/", "-", "개월"]):
            return line.strip()
    return None


def is_empty(value):
    if value is None:
        return True
    if isinstance(value, float):
        try:
            import math

            return math.isnan(value)
        except Exception:
            pass
    return str(value).strip() in ("", "nan", "None")


def normalize_text(value):
    return clean_name(format_text(str(value)))


TEACHER_ALIAS_RULES = [
    ("Ray (윤정원)", ["Ray", "윤정원", "레이"]),
]

NON_STUDENT_EXACT = {
    "일대일",
    "일대일수업",
    "중국어회화",
}

NON_STUDENT_KEYWORDS = (
    "일대일",
    "수업",
    "회화",
    "영어",
    "일본어",
    "중국어",
    "한국어",
    "토익",
    "오픽",
    "프리토킹",
    "정규반",
    "토요반",
    "입문",
    "기초",
    "초급",
    "중급",
    "고급",
    "문법",
    "과정",
    "jlpt",
    "hsk",
    "topik",
    "lv",
    "level",
)


def normalize_teacher_name(value):
    teacher = capitalize_first_word_if_english(format_text(str(value)))
    teacher_key = re.sub(r"[^0-9A-Za-z가-힣]+", "", teacher).lower()

    for canonical_name, keywords in TEACHER_ALIAS_RULES:
        keyword_keys = [
            re.sub(r"[^0-9A-Za-z가-힣]+", "", keyword).lower()
            for keyword in keywords
        ]
        if any(keyword_key and keyword_key in teacher_key for keyword_key in keyword_keys):
            return canonical_name

    return teacher


def looks_like_student_name(value):
    if is_empty(value) or isinstance(value, (int, float)):
        return False

    name = normalize_text(value)
    if not (2 <= len(name) <= 10):
        return False
    if name in NON_STUDENT_EXACT:
        return False
    lower_name = name.lower()
    if any(keyword in lower_name for keyword in NON_STUDENT_KEYWORDS):
        return False

    return re.fullmatch(r"[A-Za-z가-힣]+", name) is not None


def is_summary_row(row, time_value):
    texts = []
    for value in row.values:
        if is_empty(value) or isinstance(value, (int, float)):
            continue
        texts.append(normalize_text(value))

    if any(text.startswith("※") for text in texts):
        return True

    return is_empty(time_value) and any(text in SUMMARY_MARKERS for text in texts)


def parse_sheet(workbook_path, sheet_name, header_row, day_col_idx=None, preferred_course_col=None):
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    if sheet_name not in workbook.sheetnames:
        return []

    try:
        worksheet = workbook[sheet_name]
    except KeyError:
        return []

    df = pd.read_excel(
        workbook_path,
        header=header_row - 1,
        sheet_name=sheet_name,
        engine="openpyxl",
    )
    df.columns = [
        column if isinstance(column, (int, float)) else str(column).strip()
        for column in df.columns
    ]

    def find(keywords, exclude=()):
        for column in df.columns:
            label = str(column).strip()
            if any(keyword.lower() in label.lower() for keyword in keywords):
                if not any(ex.lower() in label.lower() for ex in exclude):
                    return column
        return None

    teacher_col = find(["강사"], exclude=["인원", "보"])
    time_col = find(["시간", "time"])

    course_col = preferred_course_col if preferred_course_col in df.columns else None
    if course_col is None:
        course_col = find(["과정"]) or find(["구분2"]) or find(["구분1"])

    if day_col_idx is not None and 0 < day_col_idx <= len(df.columns):
        day_col = df.columns[day_col_idx - 1]
    else:
        day_col = find(["요일"])

    student_cols = [
        column for column in df.columns
        if isinstance(column, (int, float)) and 1 <= column <= 20
    ]
    student_col_positions = {
        column: list(df.columns).index(column) + 1
        for column in student_cols
    }

    if not teacher_col:
        return []

    records = []
    cur = dict(teacher=None, day=None, time=None, course=None, students=[])

    def flush():
        if cur["teacher"] and cur["students"]:
            records.append({
                "강사": normalize_teacher_name(cur["teacher"]),
                "과정": format_text(str(cur["course"])) if cur["course"] else "",
                "요일": format_text(str(cur["day"])) if cur["day"] else "",
                "시간": format_text(str(cur["time"])) if cur["time"] else "",
                "학생목록": cur["students"][:],
            })

    for row_idx, row in df.iterrows():
        excel_row = header_row + 1 + row_idx

        teacher_value = row.get(teacher_col) if teacher_col else None
        day_value = row.get(day_col) if day_col else None
        time_value = row.get(time_col) if time_col else None
        course_value = row.get(course_col) if course_col else None

        if is_summary_row(row, time_value):
            flush()
            break

        teacher_new = not is_empty(teacher_value)
        day_new = not is_empty(day_value)
        time_new = not is_empty(time_value)

        if teacher_new:
            new_key = (str(teacher_value), str(day_value), str(time_value))
            old_key = (str(cur["teacher"]), str(cur["day"]), str(cur["time"]))
            if new_key != old_key:
                flush()
                cur = dict(
                    teacher=str(teacher_value),
                    day=str(day_value) if day_new else cur["day"],
                    time=str(time_value) if time_new else cur["time"],
                    course=str(course_value) if not is_empty(course_value) else cur["course"],
                    students=[],
                )
        elif day_new or time_new:
            flush()
            cur["day"] = str(day_value) if day_new else cur["day"]
            cur["time"] = str(time_value) if time_new else cur["time"]
            cur["course"] = str(course_value) if not is_empty(course_value) else cur["course"]
            cur["students"] = []

        if cur["teacher"] is None:
            continue

        for student_col in student_cols:
            value = row.get(student_col)
            if not looks_like_student_name(value):
                continue

            student_name = normalize_text(value)
            excel_col = student_col_positions[student_col]
            comment_cell = worksheet.cell(row=excel_row, column=excel_col)
            duration = extract_duration(comment_cell.comment.text if comment_cell.comment else None)

            cur["students"].append({
                "name": student_name,
                "duration": duration,
            })

    flush()
    return records


def parse_language_records(workbook_path, sheet_configs=None):
    records = []
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    available_sheets = set(workbook.sheetnames)

    for sheet_name, header_row, day_col_idx, preferred_course_col in sheet_configs or SHEET_CONFIGS:
        if sheet_name not in available_sheets:
            continue
        records.extend(
            parse_sheet(
                workbook_path,
                sheet_name,
                header_row,
                day_col_idx,
                preferred_course_col,
            )
        )
    return records
