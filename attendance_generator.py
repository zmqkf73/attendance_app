import calendar
import re
from copy import copy
from datetime import datetime
from io import BytesIO

import holidays
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break, RowBreak


def convert_non_string_to_string(value):
    return str(value) if not isinstance(value, str) else value


def remove_newlines(text):
    return text.replace("\n", " ")


def strip_edges(text):
    return text.strip()


def reduce_whitespace(text):
    return re.sub(r"\s+", " ", text)


def remove_all_whitespace(text):
    return re.sub(r"\s+", "", text)


def insert_space_before_brackets(text):
    return re.sub(r"([\w가-힣])(?=[\(\[\{])", r"\1 ", text)


def replace_tilde_with_dash(text):
    return text.replace("~", "-")


def insert_space_between_adjacent_brackets(text):
    return re.sub(r"(\))(?=\()", r"\1 ", text)


def capitalize_first_word_if_english(text):
    match = re.match(r"^([A-Za-z]+)(\s|$)", text)
    if match:
        first = match.group(1)
        rest = text[len(first):]
        return first.capitalize() + rest
    return text


def format_text(text):
    text = convert_non_string_to_string(text)
    text = remove_newlines(text)
    text = insert_space_before_brackets(text)
    text = insert_space_between_adjacent_brackets(text)
    text = replace_tilde_with_dash(text)
    text = reduce_whitespace(text)
    text = strip_edges(text)
    return text


def clean_name(name):
    name = convert_non_string_to_string(name)
    name = remove_newlines(name)
    name = strip_edges(name)
    name = remove_all_whitespace(name)
    return name


def preprocess_duration(text):
    if not text:
        return None
    text = strip_edges(text)
    text = reduce_whitespace(text)
    text = replace_tilde_with_dash(text)
    text = insert_space_before_brackets(text)
    text = insert_space_between_adjacent_brackets(text)
    return text


_CLASS_SPACING_WORDS = (
    "경우",
    "예정",
    "완료",
    "가능",
    "불가",
    "필요",
    "문의",
    "참고",
)


def polish_class_name(text):
    text = format_text(text)
    for word in _CLASS_SPACING_WORDS:
        text = re.sub(
            rf"([가-힣]{{2,}})({word})(?=$|[\s,.)/+-])",
            rf"\1 \2",
            text,
        )
    return text


_YOIL_MAP = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4, "토": 5, "일": 6}
_YOIL_ORDER = ["월", "화", "수", "목", "금", "토", "일"]


def _normalize_yoil(yoil_str):
    text = str(yoil_str)
    text = re.sub(r"[(（][^)）]*[)）]", "", text)
    text = re.sub(r"\s+", "", text)
    text = text.replace("요일", "")
    return text


def parse_weekday_indices(yoil_str):
    text = _normalize_yoil(yoil_str)
    found = [idx for char, idx in _YOIL_MAP.items() if char in text]
    return sorted(set(found)) if found else None


def format_day_display(yoil_str):
    text = _normalize_yoil(yoil_str)
    chars = [char for char in _YOIL_ORDER if char in text]
    return "".join(chars) if chars else str(yoil_str)


def shrink_font_to_fit(cell, max_chars, min_size=6.0):
    if not cell.value or not isinstance(cell.value, str):
        return
    text_len = len(cell.value)
    if text_len <= max_chars:
        return
    current_size = (cell.font.size or 11) if cell.font else 11
    new_size = max(min_size, round(current_size * max_chars / text_len, 1))
    new_font = copy(cell.font) if cell.font else Font()
    new_font.size = new_size
    cell.font = new_font


def _copy_cell(src_cell, dst_cell, copy_value=True):
    dst_cell.value = src_cell.value if copy_value else None

    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def _copy_row(template_ws, src_row, dst_ws, dst_row, max_col, copy_values=True):
    dst_ws.row_dimensions[dst_row].height = template_ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        _copy_cell(
            template_ws.cell(row=src_row, column=col),
            dst_ws.cell(row=dst_row, column=col),
            copy_value=copy_values,
        )


def _merged_range_string(min_col, min_row, max_col, max_row):
    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def _apply_template_block_merges(
    template_ws,
    dst_ws,
    block_start_row,
    block_rows,
    insertion_rel_row,
    extra_rows,
):
    for merged_range in template_ws.merged_cells.ranges:
        if merged_range.max_row > block_rows:
            continue

        row_shift = extra_rows if merged_range.min_row >= insertion_rel_row else 0
        dst_ws.merge_cells(
            _merged_range_string(
                merged_range.min_col,
                block_start_row + merged_range.min_row - 1 + row_shift,
                merged_range.max_col,
                block_start_row + merged_range.max_row - 1 + row_shift,
            )
        )


def _copy_template_block(template_ws, dst_ws, block_start_row, block_rows, max_col):
    for src_row in range(1, block_rows + 1):
        _copy_row(template_ws, src_row, dst_ws, block_start_row + src_row - 1, max_col)

    _apply_template_block_merges(
        template_ws,
        dst_ws,
        block_start_row,
        block_rows,
        block_rows + 1,
        0,
    )


def _clear_block_merges(dst_ws, block_start_row, block_end_row):
    to_unmerge = [
        str(merged_range)
        for merged_range in dst_ws.merged_cells.ranges
        if block_start_row <= merged_range.min_row and merged_range.max_row <= block_end_row
    ]
    for merged_range in to_unmerge:
        dst_ws.unmerge_cells(merged_range)


def _expand_block_for_extra_students(
    template_ws,
    dst_ws,
    block_start_row,
    block_rows,
    max_col,
    student_last_rel_row,
    extra_rows,
):
    if extra_rows <= 0:
        return

    insertion_rel_row = student_last_rel_row + 1
    insert_at_row = block_start_row + insertion_rel_row - 1
    block_end_row = block_start_row + block_rows - 1

    _clear_block_merges(dst_ws, block_start_row, block_end_row)
    dst_ws.insert_rows(insert_at_row, extra_rows)

    for extra_idx in range(extra_rows):
        dst_row = insert_at_row + extra_idx
        _copy_row(
            template_ws,
            student_last_rel_row,
            dst_ws,
            dst_row,
            max_col,
            copy_values=False,
        )

    _apply_template_block_merges(
        template_ws,
        dst_ws,
        block_start_row,
        block_rows,
        insertion_rel_row,
        extra_rows,
    )


def _build_inline_font_from_cell(cell, size):
    color = None
    if cell.font and cell.font.color and getattr(cell.font.color, "type", None) == "rgb":
        color = cell.font.color.rgb

    return InlineFont(
        rFont=(cell.font.name if cell.font else "Calibri"),
        sz=size,
        b=(cell.font.bold if cell.font else False),
        i=(cell.font.italic if cell.font else False),
        color=color,
        u=(cell.font.underline if cell.font else None),
        strike=(cell.font.strike if cell.font else False),
        vertAlign=(cell.font.vertAlign if cell.font else None),
    )


def set_class_text(cell, template_value, class_name):
    prefix = str(template_value or "")
    if prefix.endswith("ABC"):
        prefix = prefix[:-3]

    class_name = class_name or ""
    if not class_name:
        cell.value = prefix
        return

    base_size = (cell.font.size or 11) if cell.font else 11
    suffix_max_chars = 18
    suffix_size = base_size
    if len(class_name) > suffix_max_chars:
        suffix_size = max(6.0, round(base_size * suffix_max_chars / len(class_name), 1))

    prefix_font = _build_inline_font_from_cell(cell, base_size)
    class_font = _build_inline_font_from_cell(cell, suffix_size)
    cell.value = CellRichText(
        TextBlock(prefix_font, prefix),
        TextBlock(class_font, class_name),
    )


def generate_attendance(
    records,
    template_path,
    year=None,
    month=None,
    day_type="주중",
    manual_holidays=None,
    manual_includes=None,
):
    manual_holidays = set(manual_holidays or [])
    manual_includes = set(manual_includes or [])

    workbook = load_workbook(str(template_path), data_only=False, keep_links=False, keep_vba=False)
    template_ws = workbook.worksheets[0]
    template_rows = template_ws.max_row
    template_cols = template_ws.max_column
    student_slots = 11

    class_rel = teacher_rel = korean_rel = None
    class_template_value = None
    for row in template_ws.iter_rows(min_row=1, max_row=template_rows, max_col=template_cols):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            value = cell.value
            if class_rel is None and "CLASS:" in value:
                class_rel = (cell.row - 1, cell.column)
                class_template_value = value
            if teacher_rel is None and "담임" in value and "강사" in value:
                teacher_rel = (cell.row - 1, cell.column)
            if korean_rel is None and value.strip() == "Korean":
                korean_rel = (cell.row - 1, cell.column)

    if korean_rel is None:
        raise ValueError("Template does not contain a 'Korean' header cell.")

    student_start_rel_row = korean_rel[0] + 2
    student_last_rel_row = student_start_rel_row + student_slots - 1
    student_name_alignment = copy(
        template_ws.cell(row=student_start_rel_row, column=korean_rel[1]).alignment
    )

    today = datetime.today()
    used_year = year or today.year
    used_month = month or today.month

    days_kor = ["월", "화", "수", "목", "금", "토", "일"]
    _, last_day = calendar.monthrange(used_year, used_month)

    kr_holidays = holidays.KR(years=used_year)
    for holiday_date in manual_holidays:
        kr_holidays[holiday_date] = "사용자 지정 공휴일"
    for include_date in manual_includes:
        if include_date in kr_holidays:
            del kr_holidays[include_date]

    all_month_days = []
    for day in range(1, last_day + 1):
        date_obj = datetime(used_year, used_month, day)
        if date_obj.date() not in kr_holidays:
            all_month_days.append((date_obj.weekday(), date_obj.day))

    def get_valid_dates_for_record(record):
        yoil = record.get("요일", "")
        indices = parse_weekday_indices(yoil)
        if indices:
            return [(days_kor[weekday], day_num) for weekday, day_num in all_month_days if weekday in indices]
        if day_type == "토요일":
            return [(days_kor[weekday], day_num) for weekday, day_num in all_month_days if weekday == 5]
        return [(days_kor[weekday], day_num) for weekday, day_num in all_month_days if weekday < 5]

    teacher_to_records = {}
    for record in records:
        teacher = record.get("강사")
        if not isinstance(teacher, str):
            continue
        teacher = teacher.strip()
        if not teacher or teacher.lower() == "nan" or teacher == "강사":
            continue
        teacher_to_records.setdefault(teacher, []).append(record)

    teacher_block_layouts = {}
    for teacher, teacher_records in teacher_to_records.items():
        ws = workbook.copy_worksheet(template_ws)
        ws.title = teacher

        block_layouts = []
        first_extra_rows = max(0, len(teacher_records[0].get("학생목록", [])) - student_slots)
        _expand_block_for_extra_students(
            template_ws,
            ws,
            1,
            template_rows,
            template_cols,
            student_last_rel_row,
            first_extra_rows,
        )
        block_layouts.append((1, first_extra_rows))

        current_start_row = 1 + template_rows + first_extra_rows
        for record in teacher_records[1:]:
            extra_rows = max(0, len(record.get("학생목록", [])) - student_slots)
            _copy_template_block(
                template_ws,
                ws,
                current_start_row,
                template_rows,
                template_cols,
            )
            _expand_block_for_extra_students(
                template_ws,
                ws,
                current_start_row,
                template_rows,
                template_cols,
                student_last_rel_row,
                extra_rows,
            )
            block_layouts.append((current_start_row, extra_rows))
            current_start_row += template_rows + extra_rows

        teacher_block_layouts[teacher] = block_layouts

    for teacher, teacher_records in teacher_to_records.items():
        ws = workbook[teacher]
        for record, (start_row, extra_rows) in zip(teacher_records, teacher_block_layouts[teacher]):
            course = record.get("과정", "")
            class_name = course.split("/")[0] if isinstance(course, str) and course.strip() else ""
            class_name = polish_class_name(class_name)
            time_value = record.get("시간", "")
            day_value = record.get("요일", "")
            students = record.get("학생목록", [])

            ws.cell(row=start_row + 2, column=2).value = f"{str(used_year)[2:]}년 {used_month}월"

            time_cell = ws.cell(row=start_row + 2, column=7)
            time_cell.value = f"{format_day_display(day_value)} {time_value}"
            shrink_font_to_fit(time_cell, 20)

            valid_dates = get_valid_dates_for_record(record)
            for idx in range(23):
                weekday_cell = ws.cell(row=start_row + 4, column=7 + idx)
                date_cell = ws.cell(row=start_row + 5, column=7 + idx)
                if idx < len(valid_dates):
                    weekday_cell.value = valid_dates[idx][0]
                    date_cell.value = valid_dates[idx][1]
                else:
                    weekday_cell.value = None
                    date_cell.value = None

            if class_rel and class_template_value:
                class_cell = ws.cell(row=start_row + class_rel[0], column=class_rel[1])
                set_class_text(class_cell, class_template_value, class_name)

            if teacher_rel:
                teacher_cell = ws.cell(row=start_row + teacher_rel[0], column=teacher_rel[1])
                teacher_cell.value = f"담임 강사: {teacher}"
                shrink_font_to_fit(teacher_cell, 25)

            korean_col = korean_rel[1]
            student_start_row = start_row + student_start_rel_row - 1
            duration_col = korean_col + 3

            for idx, student_dict in enumerate(students):
                name = student_dict.get("name")
                if not name:
                    continue
                name_cell = ws.cell(row=student_start_row + idx, column=korean_col)
                name_cell.value = name
                name_cell.alignment = copy(student_name_alignment)
                shrink_font_to_fit(name_cell, 10)

                duration = student_dict.get("duration")
                if duration:
                    duration_cell = ws.cell(row=student_start_row + idx, column=duration_col)
                    duration_cell.value = preprocess_duration(duration)
                    duration_cell.font = Font(size=8)

            for idx in range(student_slots + extra_rows):
                ws.cell(row=student_start_row + idx, column=1).value = idx + 1

        total_rows = 0
        row_breaks = RowBreak()
        for idx, (block_start, extra_rows) in enumerate(teacher_block_layouts[teacher]):
            block_height = template_rows + extra_rows
            total_rows = max(total_rows, block_start + block_height - 1)
            if idx > 0:
                row_breaks.append(Break(id=block_start - 1))

        last_column_letter = get_column_letter(template_cols)
        ws.print_area = f"$A$1:${last_column_letter}${total_rows}"
        ws.row_breaks = row_breaks

    if "ABC" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["ABC"]

    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    return output_stream
